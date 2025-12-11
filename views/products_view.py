import customtkinter as ctk
from tkinter import messagebox
from controllers.products import obtener_productos, agregar_producto, ajustar_stock, editar_producto, eliminar_producto
from views.nuevo_producto_form_mejorado import NuevoProductoFormMejorado
import csv
import json
import pandas as pd
from datetime import datetime

class ProductsView(ctk.CTkFrame):
    def __init__(self, parent, user=None):
        super().__init__(parent, fg_color="#F5F5F5")
        self.user = user
        self.productos = []
        self.pack(fill="both", expand=True)

        # Header con título y botones
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=(10, 5))

        # Título
        title_frame = ctk.CTkFrame(header, fg_color="transparent")
        title_frame.pack(side="left")
        
        ctk.CTkLabel(
            title_frame, 
            text="📦 Productos", 
            font=("Segoe UI", 20, "bold"),
            text_color="#333333"
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            title_frame, 
            text="Gestión completa de productos y stock", 
            font=("Segoe UI", 11),
            text_color="#666666"
        ).pack(anchor="w")

        # Botones de acción
        buttons_frame = ctk.CTkFrame(header, fg_color="transparent")
        buttons_frame.pack(side="right")
        
        ctk.CTkButton(
            buttons_frame,
            text="+ Nuevo Producto",
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=40,
            width=150,
            font=("Segoe UI", 12, "bold"),
            command=self.abrir_formulario_nuevo
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            buttons_frame,
            text="📥 Importar Excel",
            fg_color="#9C27B0",
            hover_color="#7B1FA2",
            height=40,
            width=140,
            font=("Segoe UI", 12, "bold"),
            command=self.importar_desde_excel
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            buttons_frame,
            text="📊 Exportar",
            fg_color="#FF9800",
            hover_color="#F57C00",
            height=40,
            width=120,
            font=("Segoe UI", 12, "bold"),
            command=self.exportar_inventario
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            buttons_frame,
            text="🔄 Actualizar",
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=40,
            width=120,
            font=("Segoe UI", 12, "bold"),
            command=self.cargar_tabla_productos
        ).pack(side="left", padx=5)

        # Estadísticas
        self.crear_estadisticas()

        # Barra de búsqueda y filtros
        self.crear_filtros()

        # Tabla de productos
        self.tabla_frame = ctk.CTkScrollableFrame(self, fg_color="white", corner_radius=10)
        self.tabla_frame.pack(fill="both", expand=True, padx=20, pady=(5, 10))

        self.cargar_tabla_productos()
    
    def crear_estadisticas(self):
        """Crear tarjetas de estadísticas"""
        stats_frame = ctk.CTkFrame(self, fg_color="transparent")
        stats_frame.pack(fill="x", padx=20, pady=(5, 10))
        
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        # Total Productos
        card1 = ctk.CTkFrame(stats_frame, fg_color="white", corner_radius=10)
        card1.grid(row=0, column=0, padx=5, sticky="ew")
        
        icon_frame1 = ctk.CTkFrame(card1, fg_color="#E3F2FD", corner_radius=8, width=50, height=50)
        icon_frame1.pack(side="left", padx=15, pady=15)
        icon_frame1.pack_propagate(False)
        ctk.CTkLabel(icon_frame1, text="📦", font=("Segoe UI", 24)).pack(expand=True)
        
        info_frame1 = ctk.CTkFrame(card1, fg_color="transparent")
        info_frame1.pack(side="left", fill="both", expand=True, pady=15)
        ctk.CTkLabel(info_frame1, text="Total Productos", font=("Segoe UI", 10), text_color="#666666", anchor="w").pack(anchor="w")
        self.total_productos_label = ctk.CTkLabel(info_frame1, text="0", font=("Segoe UI", 20, "bold"), text_color="#333333", anchor="w")
        self.total_productos_label.pack(anchor="w")
        
        # Stock Total
        card2 = ctk.CTkFrame(stats_frame, fg_color="white", corner_radius=10)
        card2.grid(row=0, column=1, padx=5, sticky="ew")
        
        icon_frame2 = ctk.CTkFrame(card2, fg_color="#E8F5E9", corner_radius=8, width=50, height=50)
        icon_frame2.pack(side="left", padx=15, pady=15)
        icon_frame2.pack_propagate(False)
        ctk.CTkLabel(icon_frame2, text="📈", font=("Segoe UI", 24)).pack(expand=True)
        
        info_frame2 = ctk.CTkFrame(card2, fg_color="transparent")
        info_frame2.pack(side="left", fill="both", expand=True, pady=15)
        ctk.CTkLabel(info_frame2, text="Stock Total", font=("Segoe UI", 10), text_color="#666666", anchor="w").pack(anchor="w")
        self.stock_total_label = ctk.CTkLabel(info_frame2, text="0 unid.", font=("Segoe UI", 20, "bold"), text_color="#4CAF50", anchor="w")
        self.stock_total_label.pack(anchor="w")
        
        # Stock Bajo
        card3 = ctk.CTkFrame(stats_frame, fg_color="white", corner_radius=10)
        card3.grid(row=0, column=2, padx=5, sticky="ew")
        
        icon_frame3 = ctk.CTkFrame(card3, fg_color="#FFEBEE", corner_radius=8, width=50, height=50)
        icon_frame3.pack(side="left", padx=15, pady=15)
        icon_frame3.pack_propagate(False)
        ctk.CTkLabel(icon_frame3, text="⚠️", font=("Segoe UI", 24)).pack(expand=True)
        
        info_frame3 = ctk.CTkFrame(card3, fg_color="transparent")
        info_frame3.pack(side="left", fill="both", expand=True, pady=15)
        ctk.CTkLabel(info_frame3, text="Stock Bajo", font=("Segoe UI", 10), text_color="#666666", anchor="w").pack(anchor="w")
        self.stock_bajo_label = ctk.CTkLabel(info_frame3, text="0", font=("Segoe UI", 20, "bold"), text_color="#E53935", anchor="w")
        self.stock_bajo_label.pack(anchor="w")
        
        # Valor Total
        card4 = ctk.CTkFrame(stats_frame, fg_color="white", corner_radius=10)
        card4.grid(row=0, column=3, padx=5, sticky="ew")
        
        icon_frame4 = ctk.CTkFrame(card4, fg_color="#FFF3E0", corner_radius=8, width=50, height=50)
        icon_frame4.pack(side="left", padx=15, pady=15)
        icon_frame4.pack_propagate(False)
        ctk.CTkLabel(icon_frame4, text="💰", font=("Segoe UI", 24)).pack(expand=True)
        
        info_frame4 = ctk.CTkFrame(card4, fg_color="transparent")
        info_frame4.pack(side="left", fill="both", expand=True, pady=15)
        ctk.CTkLabel(info_frame4, text="Valor Total", font=("Segoe UI", 10), text_color="#666666", anchor="w").pack(anchor="w")
        self.valor_total_label = ctk.CTkLabel(info_frame4, text="$0.00", font=("Segoe UI", 20, "bold"), text_color="#FF9800", anchor="w")
        self.valor_total_label.pack(anchor="w")
    
    def crear_filtros(self):
        """Crear barra de búsqueda y filtros"""
        filtros_frame = ctk.CTkFrame(self, fg_color="transparent")
        filtros_frame.pack(fill="x", padx=20, pady=(10, 5))
        
        # Búsqueda
        search_frame = ctk.CTkFrame(filtros_frame, fg_color="white", corner_radius=10, height=45)
        search_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))
        search_frame.pack_propagate(False)
        
        ctk.CTkLabel(search_frame, text="🔍", font=("Segoe UI", 16)).pack(side="left", padx=10)
        self.search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="Buscar producto por nombre o código...",
            border_width=0,
            fg_color="white",
            font=("Segoe UI", 11)
        )
        self.search_entry.pack(side="left", fill="both", expand=True, padx=(0, 10))
        self.search_entry.bind("<KeyRelease>", lambda e: self.filtrar_productos())
        
        # Filtro de nivel de stock
        self.nivel_combo = ctk.CTkComboBox(
            filtros_frame,
            values=["Todos los niveles", "Stock bajo", "Stock medio", "Stock alto"],
            width=180,
            height=45,
            corner_radius=10,
            font=("Segoe UI", 11),
            command=lambda e: self.filtrar_productos()
        )
        self.nivel_combo.set("Todos los niveles")
        self.nivel_combo.pack(side="left", padx=5)

    def cargar_tabla_productos(self):
        """Carga o recarga los productos en la tabla."""
        # Limpiar el frame de la tabla
        for widget in self.tabla_frame.winfo_children():
            widget.destroy()

        self.productos = obtener_productos()
        
        # Actualizar estadísticas
        self.actualizar_estadisticas()

        # Header de la tabla
        header_frame = ctk.CTkFrame(self.tabla_frame, fg_color="#F5F5F5")
        header_frame.pack(fill="x", pady=(0, 5))

        headers = [("Código", 150), ("Nombre", 350), ("Precio", 130), ("Stock", 130), ("Valor", 150), ("Acciones", 180)]
        
        for header, width in headers:
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=("Segoe UI", 13, "bold"),
                text_color="#333333",
                width=width
            ).pack(side="left", padx=8)

        # Filas de productos
        for producto in self.productos:
            self.crear_fila_producto(producto)
    
    def actualizar_estadisticas(self):
        """Actualizar las tarjetas de estadísticas"""
        total = len(self.productos)
        stock_total = sum(p.get("stock", 0) for p in self.productos)
        stock_bajo = sum(1 for p in self.productos if p.get("stock", 0) < 10)
        valor_total = sum(float(p.get("precio", 0)) * int(p.get("stock", 0)) for p in self.productos)
        
        self.total_productos_label.configure(text=str(total))
        self.stock_total_label.configure(text=f"{stock_total} unid.")
        self.stock_bajo_label.configure(text=str(stock_bajo))
        self.valor_total_label.configure(text=f"${valor_total:,.2f}")
    
    def crear_fila_producto(self, producto):
        """Crear fila de producto"""
        fila = ctk.CTkFrame(self.tabla_frame, fg_color="white", height=38)
        fila.pack(fill="x", pady=1)
        fila.pack_propagate(False)
        
        # Código
        codigo = producto.get("codigo", producto.get("id_producto", "N/A"))
        ctk.CTkLabel(fila, text=str(codigo), font=("Segoe UI", 12), text_color="#666666", width=150).pack(side="left", padx=8)
        
        # Nombre
        nombre = producto.get("nombre", "N/A")
        ctk.CTkLabel(fila, text=nombre[:45], font=("Segoe UI", 12), text_color="#333333", width=350, anchor="w").pack(side="left", padx=8)
        
        # Precio
        precio = float(producto.get("precio", 0))
        ctk.CTkLabel(fila, text=f"${precio:.2f}", font=("Segoe UI", 12, "bold"), text_color="#333333", width=130).pack(side="left", padx=8)
        
        # Stock con color
        stock = int(producto.get("stock", 0))
        color = "#E53935" if stock < 10 else "#FF9800" if stock < 20 else "#4CAF50"
        ctk.CTkLabel(fila, text=f"{stock} unid.", font=("Segoe UI", 12, "bold"), text_color=color, width=130).pack(side="left", padx=8)
        
        # Valor total
        valor = precio * stock
        ctk.CTkLabel(fila, text=f"${valor:,.2f}", font=("Segoe UI", 12), text_color="#333333", width=150).pack(side="left", padx=8)
        
        # Botones de acción
        acciones_frame = ctk.CTkFrame(fila, fg_color="transparent")
        acciones_frame.pack(side="left", padx=5)
        
        # Botón Editar
        ctk.CTkButton(
            acciones_frame,
            text="✏️",
            fg_color="transparent",
            text_color="#FF9800",
            hover_color="#FFF3E0",
            width=40,
            height=35,
            font=("Segoe UI", 16),
            command=lambda p=producto: self.editar_producto(p)
        ).pack(side="left", padx=2)
        
        # Botón Ajustar
        ctk.CTkButton(
            acciones_frame,
            text="⚙",
            fg_color="transparent",
            text_color="#2196F3",
            hover_color="#E3F2FD",
            width=40,
            height=35,
            font=("Segoe UI", 16),
            command=lambda p=producto: self.ajustar_stock(p)
        ).pack(side="left", padx=2)
        
        # Botón Eliminar
        ctk.CTkButton(
            acciones_frame,
            text="🗑️",
            fg_color="transparent",
            text_color="#F44336",
            hover_color="#FFEBEE",
            width=40,
            height=35,
            font=("Segoe UI", 16),
            command=lambda p=producto: self.eliminar_producto(p)
        ).pack(side="left", padx=2)
    
    def ajustar_stock(self, producto):
        """Ajustar stock de un producto con diálogo mejorado"""
        # Crear ventana de diálogo personalizada
        dialog = ctk.CTkToplevel(self)
        dialog.title("Ajustar Stock")
        dialog.geometry("400x350")
        dialog.transient(self)
        dialog.grab_set()
        
        # Centrar ventana
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (350 // 2)
        dialog.geometry(f"400x350+{x}+{y}")
        
        # Contenido
        main_frame = ctk.CTkFrame(dialog, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            main_frame,
            text="Ajustar Stock",
            font=("Segoe UI", 20, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 10))
        
        ctk.CTkLabel(
            main_frame,
            text=producto.get("nombre", "Producto"),
            font=("Segoe UI", 14),
            text_color="#666666"
        ).pack(pady=(0, 20))
        
        # Stock actual
        stock_actual = int(producto.get("stock", 0))
        ctk.CTkLabel(
            main_frame,
            text=f"Stock actual: {stock_actual} unidades",
            font=("Segoe UI", 12, "bold"),
            text_color="#E91E63"
        ).pack(pady=(0, 20))
        
        # Tipo de ajuste
        ctk.CTkLabel(
            main_frame,
            text="Tipo de ajuste",
            font=("Segoe UI", 12),
            text_color="#666666",
            anchor="w"
        ).pack(fill="x", pady=(0, 5))
        
        tipo_var = ctk.StringVar(value="Entrada")
        tipo_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tipo_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkRadioButton(
            tipo_frame,
            text="Entrada (agregar)",
            variable=tipo_var,
            value="Entrada",
            font=("Segoe UI", 11)
        ).pack(side="left", padx=10)
        
        ctk.CTkRadioButton(
            tipo_frame,
            text="Salida (quitar)",
            variable=tipo_var,
            value="Salida",
            font=("Segoe UI", 11)
        ).pack(side="left", padx=10)
        
        # Cantidad
        ctk.CTkLabel(
            main_frame,
            text="Cantidad",
            font=("Segoe UI", 12),
            text_color="#666666",
            anchor="w"
        ).pack(fill="x", pady=(0, 5))
        
        cantidad_entry = ctk.CTkEntry(
            main_frame,
            placeholder_text="Ingresa la cantidad",
            height=40,
            font=("Segoe UI", 12)
        )
        cantidad_entry.pack(fill="x", pady=(0, 20))
        
        # Botones
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(10, 0))
        
        def aplicar_ajuste():
            try:
                cantidad = int(cantidad_entry.get())
                if cantidad <= 0:
                    messagebox.showwarning("Cantidad inválida", "La cantidad debe ser mayor a 0")
                    return
                
                tipo = tipo_var.get()
                
                # Validar que no se quite más stock del disponible
                if tipo == "Salida" and cantidad > stock_actual:
                    messagebox.showwarning("Stock insuficiente", f"No puedes quitar {cantidad} unidades. Stock actual: {stock_actual}")
                    return
                
                # Aplicar ajuste
                if ajustar_stock(producto["id_producto"], cantidad, tipo):
                    messagebox.showinfo("Éxito", f"Stock ajustado correctamente\n\nTipo: {tipo}\nCantidad: {cantidad}")
                    dialog.destroy()
                    self.cargar_tabla_productos()
                else:
                    messagebox.showerror("Error", "No se pudo ajustar el stock")
            
            except ValueError:
                messagebox.showwarning("Cantidad inválida", "Ingresa un número válido")
        
        ctk.CTkButton(
            btn_frame,
            text="Aplicar",
            fg_color="#E91E63",
            hover_color="#C2185B",
            height=40,
            font=("Segoe UI", 12, "bold"),
            command=aplicar_ajuste
        ).pack(side="left", expand=True, fill="x", padx=(0, 5))
        
        ctk.CTkButton(
            btn_frame,
            text="Cancelar",
            fg_color="white",
            text_color="#666666",
            border_width=2,
            border_color="#E0E0E0",
            hover_color="#F5F5F5",
            height=40,
            font=("Segoe UI", 12),
            command=dialog.destroy
        ).pack(side="left", expand=True, fill="x", padx=(5, 0))
    
    def filtrar_productos(self):
        """Filtrar productos por búsqueda y nivel de stock"""
        termino = self.search_entry.get().lower()
        nivel = self.nivel_combo.get()
        
        # Limpiar tabla
        for widget in self.tabla_frame.winfo_children():
            if isinstance(widget, ctk.CTkFrame) and widget != self.tabla_frame.winfo_children()[0]:
                widget.destroy()
        
        # Filtrar por búsqueda
        productos_filtrados = [
            p for p in self.productos
            if termino in p.get("nombre", "").lower() or
               termino in str(p.get("codigo", "")).lower() or
               termino in str(p.get("codigo_barras", "")).lower() or
               termino in str(p.get("id_producto", "")).lower()
        ]
        
        # Filtrar por nivel de stock
        if nivel == "Stock bajo":
            productos_filtrados = [p for p in productos_filtrados if p.get("stock", 0) < 10]
        elif nivel == "Stock medio":
            productos_filtrados = [p for p in productos_filtrados if 10 <= p.get("stock", 0) < 20]
        elif nivel == "Stock alto":
            productos_filtrados = [p for p in productos_filtrados if p.get("stock", 0) >= 20]
        
        for producto in productos_filtrados:
            self.crear_fila_producto(producto)
    
    def exportar_inventario(self):
        """Mostrar opciones de exportación"""
        # Crear ventana de opciones
        export_dialog = ctk.CTkToplevel(self)
        export_dialog.title("Exportar Inventario")
        export_dialog.geometry("450x400")
        export_dialog.transient(self)
        export_dialog.grab_set()
        
        # Centrar ventana
        export_dialog.update_idletasks()
        x = (export_dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (export_dialog.winfo_screenheight() // 2) - (400 // 2)
        export_dialog.geometry(f"450x400+{x}+{y}")
        
        # Contenido
        main_frame = ctk.CTkFrame(export_dialog, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            main_frame,
            text="📊 Exportar Inventario",
            font=("Segoe UI", 18, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 10))
        
        ctk.CTkLabel(
            main_frame,
            text=f"Total de productos: {len(self.productos)}",
            font=("Segoe UI", 12),
            text_color="#666666"
        ).pack(pady=(0, 20))
        
        # Opciones de exportación mejoradas
        ctk.CTkButton(
            main_frame,
            text="📊 Exportar a Excel (.xlsx)",
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=50,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.exportar_excel_pandas(export_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(
            main_frame,
            text="📄 Exportar a CSV",
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=50,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.exportar_csv_pandas(export_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(
            main_frame,
            text="📝 Exportar a HTML",
            fg_color="#FF9800",
            hover_color="#F57C00",
            height=50,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.exportar_html_pandas(export_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(
            main_frame,
            text="📋 Exportar a JSON",
            fg_color="#9C27B0",
            hover_color="#7B1FA2",
            height=50,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.exportar_json_pandas(export_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(
            main_frame,
            text="📄 Exportar a PDF",
            fg_color="#E53935",
            hover_color="#C62828",
            height=50,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.exportar_csv_pandas(export_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(
            main_frame,
            text="Cancelar",
            fg_color="#E0E0E0",
            text_color="#666666",
            hover_color="#D0D0D0",
            height=40,
            font=("Segoe UI", 11),
            command=export_dialog.destroy
        ).pack(fill="x", pady=(10, 0))
    
    def exportar_excel_pandas(self, dialog):
        """Exportar a Excel usando Pandas"""
        try:
            from tkinter import filedialog
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"inventario_{fecha}.xlsx",
                title="Guardar Inventario como Excel"
            )
            
            if not filename:
                return

            df = pd.DataFrame(self.productos)
            # Renombrar columnas para mejor presentación
            columnas = {
                'codigo': 'Código',
                'codigo_barras': 'Código de Barras',
                'nombre': 'Nombre',
                'precio': 'Precio',
                'stock': 'Stock',
                'categoria': 'Categoría',
                'descripcion': 'Descripción'
            }
            
            # Asegurar que existan las columnas aunque estén vacías
            for col in columnas.keys():
                if col not in df.columns:
                    df[col] = ''
            
            df = df.rename(columns=columnas)
            
            # Calcular valor total
            df['Valor Total'] = df['Precio'].astype(float) * df['Stock'].astype(int)
            
            # Seleccionar columnas a exportar
            cols_exportar = ['Código', 'Código de Barras', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Valor Total', 'Descripción']
            df = df[cols_exportar]
            
            # Guardar
            df.to_excel(filename, index=False)
            
            dialog.destroy()
            messagebox.showinfo("Éxito", f"Archivo guardado correctamente:\n{filename}")
            
            # Intentar abrir el archivo
            try:
                os.startfile(filename)
            except:
                pass
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a Excel: {e}")

    def exportar_csv_pandas(self, dialog):
        """Exportar a CSV usando Pandas"""
        try:
            from tkinter import filedialog
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                initialfile=f"inventario_{fecha}.csv",
                title="Guardar Inventario como CSV"
            )
            
            if not filename:
                return

            df = pd.DataFrame(self.productos)
            df.to_csv(filename, index=False, encoding='utf-8-sig')
            
            dialog.destroy()
            messagebox.showinfo("Éxito", f"Archivo guardado correctamente:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a CSV: {e}")

    def exportar_html_pandas(self, dialog):
        """Exportar a HTML usando Pandas"""
        try:
            from tkinter import filedialog
            import webbrowser
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                defaultextension=".html",
                filetypes=[("HTML files", "*.html")],
                initialfile=f"inventario_{fecha}.html",
                title="Guardar Inventario como HTML"
            )
            
            if not filename:
                return

            df = pd.DataFrame(self.productos)
            
            # Estilizar HTML
            html = df.to_html(index=False, classes='table table-striped', border=0)
            
            html_template = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Inventario</title>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
                <style>
                    body {{ padding: 20px; }}
                    h1 {{ color: #E91E63; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <h1>Inventario Janet Rosa Bici</h1>
                <p>Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
                {html}
            </body>
            </html>
            """
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_template)
            
            dialog.destroy()
            webbrowser.open('file://' + os.path.abspath(filename))
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a HTML: {e}")

    def exportar_json_pandas(self, dialog):
        """Exportar a JSON usando Pandas"""
        try:
            from tkinter import filedialog
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json")],
                initialfile=f"inventario_{fecha}.json",
                title="Guardar Inventario como JSON"
            )
            
            if not filename:
                return

            df = pd.DataFrame(self.productos)
            df.to_json(filename, orient='records', indent=4, force_ascii=False)
            
            dialog.destroy()
            messagebox.showinfo("Éxito", f"Archivo guardado correctamente:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a JSON: {e}")

    def exportar_csv(self, dialog):
        """Exportar inventario a CSV"""
        try:
            # Generar nombre de archivo con fecha
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"inventario_{fecha}.csv"
            
            # Crear archivo CSV
            with open(filename, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                
                # Escribir encabezados
                writer.writerow(['Código', 'Código Barras', 'Nombre', 'Precio', 'Stock', 'Valor Total'])
                
                # Escribir datos
                for producto in self.productos:
                    codigo = producto.get('codigo', producto.get('id_producto', 'N/A'))
                    codigo_barras = producto.get('codigo_barras', 'N/A')
                    nombre = producto.get('nombre', 'N/A')
                    precio = float(producto.get('precio', 0))
                    stock = int(producto.get('stock', 0))
                    valor_total = precio * stock
                    
                    writer.writerow([codigo, codigo_barras, nombre, f"${precio:.2f}", stock, f"${valor_total:.2f}"])
            
            dialog.destroy()
            messagebox.showinfo("Exportación exitosa", f"Inventario exportado a:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el inventario: {str(e)}")
    
    def exportar_html(self, dialog):
        """Exportar inventario a HTML (compatible con Word)"""
        try:
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"inventario_{fecha}.html"
            
            # Calcular totales
            total_productos = len(self.productos)
            stock_total = sum(p.get("stock", 0) for p in self.productos)
            valor_total = sum(float(p.get("precio", 0)) * int(p.get("stock", 0)) for p in self.productos)
            
            # Generar filas de productos
            filas_html = ""
            for i, producto in enumerate(self.productos, 1):
                codigo = producto.get('codigo', producto.get('id_producto', 'N/A'))
                nombre = producto.get('nombre', 'N/A')
                precio = float(producto.get('precio', 0))
                stock = int(producto.get('stock', 0))
                valor = precio * stock
                
                # Color de fila alternado
                bg_color = "#F9F9F9" if i % 2 == 0 else "#FFFFFF"
                
                filas_html += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 10px; border: 1px solid #ddd;">{codigo}</td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{nombre}</td>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">${precio:,.2f}</td>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{stock}</td>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: right; font-weight: bold;">${valor:,.2f}</td>
                </tr>
                """
            
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Inventario - Janet Rosa Bici</title>
                <style>
                    body {{
                        font-family: 'Segoe UI', Arial, sans-serif;
                        margin: 40px;
                        background: white;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 30px;
                        border-bottom: 3px solid #E91E63;
                        padding-bottom: 20px;
                    }}
                    .logo {{
                        font-size: 40px;
                        margin-bottom: 10px;
                    }}
                    .business-name {{
                        font-size: 28px;
                        font-weight: bold;
                        margin-bottom: 5px;
                    }}
                    .rosa {{
                        color: #E91E63;
                    }}
                    .subtitle {{
                        font-size: 14px;
                        color: #666;
                    }}
                    .info {{
                        background: #F5F5F5;
                        padding: 15px;
                        border-radius: 8px;
                        margin-bottom: 20px;
                    }}
                    .info-row {{
                        display: flex;
                        justify-content: space-between;
                        margin: 5px 0;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 20px 0;
                    }}
                    th {{
                        background: #E91E63;
                        color: white;
                        padding: 12px;
                        text-align: left;
                        border: 1px solid #C2185B;
                    }}
                    .summary {{
                        background: #FFF0F5;
                        padding: 20px;
                        border-radius: 8px;
                        margin-top: 20px;
                    }}
                    .summary-row {{
                        display: flex;
                        justify-content: space-between;
                        margin: 10px 0;
                        font-size: 16px;
                    }}
                    .summary-total {{
                        font-size: 20px;
                        font-weight: bold;
                        color: #E91E63;
                        border-top: 2px solid #E91E63;
                        padding-top: 10px;
                        margin-top: 10px;
                    }}
                    .footer {{
                        text-align: center;
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 2px solid #E0E0E0;
                        color: #999;
                        font-size: 12px;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="logo">🚲</div>
                    <div class="business-name">
                        Janet <span class="rosa">Rosa</span> Bici
                    </div>
                    <div class="subtitle">Reporte de Inventario</div>
                </div>
                
                <div class="info">
                    <div class="info-row">
                        <strong>Fecha de generación:</strong>
                        <span>{datetime.now().strftime("%d/%m/%Y %H:%M")}</span>
                    </div>
                    <div class="info-row">
                        <strong>Total de productos:</strong>
                        <span>{total_productos}</span>
                    </div>
                    <div class="info-row">
                        <strong>Stock total:</strong>
                        <span>{stock_total} unidades</span>
                    </div>
                </div>
                
                <table>
                    <thead>
                        <tr>
                            <th>Código</th>
                            <th>Nombre del Producto</th>
                            <th style="text-align: right;">Precio</th>
                            <th style="text-align: center;">Stock</th>
                            <th style="text-align: right;">Valor Total</th>
                        </tr>
                    </thead>
                    <tbody>
                        {filas_html}
                    </tbody>
                </table>
                
                <div class="summary">
                    <div class="summary-row">
                        <span>Total de productos:</span>
                        <strong>{total_productos}</strong>
                    </div>
                    <div class="summary-row">
                        <span>Stock total:</span>
                        <strong>{stock_total} unidades</strong>
                    </div>
                    <div class="summary-row summary-total">
                        <span>Valor total del inventario:</span>
                        <strong>${valor_total:,.2f}</strong>
                    </div>
                </div>
                
                <div class="footer">
                    <p>Janet Rosa Bici - Sistema de Gestión de Inventario</p>
                    <p>Generado el {datetime.now().strftime("%d de %B de %Y a las %H:%M")}</p>
                </div>
            </body>
            </html>
            """
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Abrir en navegador
            import webbrowser
            import os
            webbrowser.open('file://' + os.path.abspath(filename))
            
            dialog.destroy()
            messagebox.showinfo(
                "Exportación exitosa",
                f"Inventario exportado a HTML:\n{filename}\n\n"
                f"Se abrió en tu navegador.\n"
                f"Para abrir en Word:\n"
                f"1. Haz clic derecho en el archivo\n"
                f"2. Abrir con > Microsoft Word"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {str(e)}")
    
    def exportar_pdf(self, dialog):
        """Exportar inventario a PDF"""
        from tkinter import filedialog
        import os
        
        try:
            # Solicitar ubicación para guardar
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"inventario_{fecha}.pdf",
                title="Guardar Inventario como PDF"
            )
            
            if not filename:
                return
            
            # Calcular totales
            total_productos = len(self.productos)
            stock_total = sum(p.get("stock", 0) for p in self.productos)
            valor_total = sum(float(p.get("precio", 0)) * int(p.get("stock", 0)) for p in self.productos)
            
            # Generar filas HTML
            filas_html = ""
            for i, producto in enumerate(self.productos, 1):
                codigo = producto.get('codigo', producto.get('id_producto', 'N/A'))
                nombre = producto.get('nombre', 'N/A')
                precio = float(producto.get('precio', 0))
                stock = int(producto.get('stock', 0))
                valor = precio * stock
                
                bg_color = "#F9F9F9" if i % 2 == 0 else "#FFFFFF"
                
                filas_html += f"""
                <tr style="background-color: {bg_color};">
                    <td style="padding: 8px; border: 1px solid #ddd; font-size: 10px;">{codigo}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; font-size: 10px;">{nombre}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-size: 10px;">${precio:,.2f}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-size: 10px;">{stock}</td>
                    <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold; font-size: 10px;">${valor:,.2f}</td>
                </tr>
                """
            
            # HTML optimizado para PDF
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Inventario PDF</title>
                <style>
                    @page {{ margin: 1cm; }}
                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; font-size: 10px; }}
                    .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #E91E63; padding-bottom: 15px; }}
                    .business-name {{ font-size: 20px; font-weight: bold; }}
                    .rosa {{ color: #E91E63; }}
                    table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
                    th {{ background: #E91E63; color: white; padding: 8px; font-size: 10px; border: 1px solid #C2185B; }}
                    .summary {{ background: #FFF0F5; padding: 15px; margin-top: 15px; }}
                    .summary-total {{ font-size: 14px; font-weight: bold; color: #E91E63; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="business-name">Janet <span class="rosa">Rosa</span> Bici</div>
                    <div>Reporte de Inventario - {datetime.now().strftime("%d/%m/%Y")}</div>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>Código</th>
                            <th>Producto</th>
                            <th>Precio</th>
                            <th>Stock</th>
                            <th>Valor</th>
                        </tr>
                    </thead>
                    <tbody>{filas_html}</tbody>
                </table>
                <div class="summary">
                    <div>Total productos: {total_productos} | Stock total: {stock_total} unidades</div>
                    <div class="summary-total">Valor total: ${valor_total:,.2f}</div>
                </div>
            </body>
            </html>
            """
            
            # Guardar HTML y abrir en navegador para imprimir como PDF
            html_filename = filename.replace('.pdf', '.html')
            with open(html_filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            import webbrowser
            webbrowser.open('file://' + os.path.abspath(html_filename))
            
            dialog.destroy()
            messagebox.showinfo(
                "✅ Inventario Abierto",
                f"Se abrió el inventario en tu navegador.\n\n"
                f"Para guardar como PDF:\n"
                f"  1. Presiona Ctrl+P\n"
                f"  2. Selecciona 'Guardar como PDF'\n"
                f"  3. Guarda el archivo\n\n"
                f"📄 Ubicación: {os.path.basename(html_filename)}"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {str(e)}")
    
    def importar_desde_excel(self):
        """Importar productos desde archivo Excel"""
        from tkinter import filedialog
        
        # Crear ventana de ayuda
        help_dialog = ctk.CTkToplevel(self)
        help_dialog.title("Importar Productos desde Excel")
        help_dialog.geometry("700x600")
        help_dialog.transient(self)
        help_dialog.grab_set()
        
        # Centrar ventana
        help_dialog.update_idletasks()
        x = (help_dialog.winfo_screenwidth() // 2) - (700 // 2)
        y = (help_dialog.winfo_screenheight() // 2) - (600 // 2)
        help_dialog.geometry(f"700x600+{x}+{y}")
        
        # Contenido
        main_frame = ctk.CTkFrame(help_dialog, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        ctk.CTkLabel(
            main_frame,
            text="📥 Importar Productos desde Excel",
            font=("Segoe UI", 20, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 20))
        
        # Instrucciones
        instrucciones_frame = ctk.CTkFrame(main_frame, fg_color="#E3F2FD", corner_radius=10)
        instrucciones_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(
            instrucciones_frame,
            text="📋 Formato del archivo Excel:",
            font=("Segoe UI", 14, "bold"),
            text_color="#1976D2",
            anchor="w"
        ).pack(anchor="w", padx=15, pady=(15, 10))
        
        instrucciones_text = """
El archivo Excel debe tener las siguientes columnas (en este orden):

1. Código (opcional) - Código del producto
2. Código de Barras (opcional) - Código de barras
3. Nombre (obligatorio) - Nombre del producto
4. Precio (obligatorio) - Precio de venta (número)
5. Stock (obligatorio) - Cantidad en inventario (número)

Ejemplo:
┌──────────┬────────────────┬─────────────────┬─────────┬────────┐
│ Código   │ Código Barras  │ Nombre          │ Precio  │ Stock  │
├──────────┼────────────────┼─────────────────┼─────────┼────────┤
│ PROD001  │ 7501234567890  │ Blusa Rosa      │ 299.00  │ 15     │
│ PROD002  │ 7501234567891  │ Pantalón Negro  │ 450.00  │ 8      │
│ PROD003  │                │ Falda Azul      │ 350.00  │ 12     │
└──────────┴────────────────┴─────────────────┴─────────┴────────┘

Notas importantes:
• La primera fila debe contener los encabezados
• Los campos Nombre, Precio y Stock son obligatorios
• El precio debe ser un número (sin símbolos)
• El stock debe ser un número entero
• Puedes dejar vacíos Código y Código de Barras
        """
        
        ctk.CTkLabel(
            instrucciones_frame,
            text=instrucciones_text,
            font=("Consolas", 10),
            text_color="#333333",
            anchor="w",
            justify="left"
        ).pack(anchor="w", padx=15, pady=(0, 15))
        
        # Botones
        buttons_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        # Botón descargar plantilla
        ctk.CTkButton(
            buttons_frame,
            text="📄 Descargar Plantilla Excel",
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=45,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.descargar_plantilla_excel(help_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        # Botón seleccionar archivo
        ctk.CTkButton(
            buttons_frame,
            text="📂 Seleccionar Archivo Excel",
            fg_color="#9C27B0",
            hover_color="#7B1FA2",
            height=45,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.procesar_excel(help_dialog)
        ).pack(fill="x", pady=(0, 10))
        
        # Botón cancelar
        ctk.CTkButton(
            buttons_frame,
            text="Cancelar",
            fg_color="#E0E0E0",
            text_color="#666666",
            hover_color="#D0D0D0",
            height=40,
            font=("Segoe UI", 11),
            command=help_dialog.destroy
        ).pack(fill="x")
    
    def descargar_plantilla_excel(self, dialog):
        """Descargar plantilla de Excel para importar productos"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos"
            
            # Encabezados
            headers = ['Código', 'Código de Barras', 'Nombre', 'Precio', 'Stock']
            ws.append(headers)
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos de ejemplo
            ejemplos = [
                ['PROD001', '7501234567890', 'Blusa Rosa', 299.00, 15],
                ['PROD002', '7501234567891', 'Pantalón Negro', 450.00, 8],
                ['PROD003', '', 'Falda Azul', 350.00, 12],
            ]
            
            for ejemplo in ejemplos:
                ws.append(ejemplo)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            
            # Guardar archivo
            filename = f"plantilla_productos_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo(
                "Plantilla creada",
                f"Plantilla descargada exitosamente:\n{filename}\n\n"
                f"Edita este archivo con tus productos y luego impórtalo."
            )
            
        except ImportError:
            # Si no está instalado openpyxl, crear CSV
            try:
                filename = f"plantilla_productos_{datetime.now().strftime('%Y%m%d')}.csv"
                
                with open(filename, 'w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    writer.writerow(['Código', 'Código de Barras', 'Nombre', 'Precio', 'Stock'])
                    writer.writerow(['PROD001', '7501234567890', 'Blusa Rosa', '299.00', '15'])
                    writer.writerow(['PROD002', '7501234567891', 'Pantalón Negro', '450.00', '8'])
                    writer.writerow(['PROD003', '', 'Falda Azul', '350.00', '12'])
                
                messagebox.showinfo(
                    "Plantilla creada (CSV)",
                    f"Plantilla descargada como CSV:\n{filename}\n\n"
                    f"Puedes abrirla con Excel y guardarla como .xlsx\n"
                    f"Edita este archivo con tus productos y luego impórtalo."
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear la plantilla: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear la plantilla: {str(e)}")
    
    def procesar_excel(self, dialog):
        """Procesar archivo Excel y agregar productos"""
        from tkinter import filedialog
        
        # Seleccionar archivo
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Archivos CSV", "*.csv"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if not filename:
            return
        
        try:
            productos_importados = []
            
            # Intentar leer con openpyxl (Excel)
            if filename.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    
                    wb = openpyxl.load_workbook(filename)
                    ws = wb.active
                    
                    # Leer encabezados para detectar columnas
                    headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
                    
                    # Detectar índices de columnas
                    idx_nombre = next((i for i, h in enumerate(headers) if 'nombre' in h), None)
                    idx_precio = next((i for i, h in enumerate(headers) if 'precio' in h), None)
                    idx_stock = next((i for i, h in enumerate(headers) if 'cantidad' in h or 'stock' in h), None)
                    idx_codigo = next((i for i, h in enumerate(headers) if 'codigo' in h and 'barras' not in h), None)
                    idx_codigo_barras = next((i for i, h in enumerate(headers) if 'barras' in h or 'codigo de barras' in h), None)
                    
                    if idx_nombre is None or idx_precio is None or idx_stock is None:
                        messagebox.showerror(
                            "Error de formato",
                            "El archivo debe tener al menos las columnas:\n"
                            "- Nombre (o nombre)\n"
                            "- Precio (o precio)\n"
                            "- Cantidad o Stock (o cantidad/stock)"
                        )
                        return
                    
                    # Leer datos
                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    
                    for row in rows:
                        if not row or not any(row):  # Saltar filas vacías
                            continue
                        
                        try:
                            nombre = str(row[idx_nombre]) if row[idx_nombre] else ''
                            precio = float(str(row[idx_precio]).replace('$', '').replace(',', '')) if row[idx_precio] else 0
                            stock = int(row[idx_stock]) if row[idx_stock] else 0
                            codigo = str(row[idx_codigo]) if idx_codigo is not None and len(row) > idx_codigo and row[idx_codigo] else ''
                            codigo_barras = str(row[idx_codigo_barras]) if idx_codigo_barras is not None and len(row) > idx_codigo_barras and row[idx_codigo_barras] else ''
                            
                            # Agregar a la lista si tiene datos válidos
                            if nombre and precio > 0:
                                productos_importados.append({
                                    'codigo': codigo,
                                    'codigo_barras': codigo_barras,
                                    'nombre': nombre,
                                    'precio': precio,
                                    'stock': stock
                                })
                        except (ValueError, IndexError, TypeError) as e:
                            print(f"Error procesando fila: {e}")
                            continue
                        
                        if nombre and precio > 0:
                            productos_importados.append({
                                'codigo': str(codigo),
                                'codigo_barras': str(codigo_barras),
                                'nombre': str(nombre),
                                'precio': precio,
                                'stock': stock
                            })
                
                except ImportError:
                    messagebox.showerror(
                        "Librería no instalada",
                        "Para importar archivos Excel (.xlsx), necesitas instalar openpyxl:\n\n"
                        "pip install openpyxl\n\n"
                        "O guarda tu archivo como CSV e intenta de nuevo."
                    )
                    return
            
            # Leer CSV
            elif filename.endswith('.csv'):
                with open(filename, 'r', encoding='utf-8') as file:
                    reader = csv.reader(file)
                    next(reader)  # Saltar encabezados
                    
                    for row in reader:
                        if not row or not any(row):
                            continue
                        
                        codigo = row[0] if len(row) > 0 else ''
                        codigo_barras = row[1] if len(row) > 1 else ''
                        nombre = row[2] if len(row) > 2 else ''
                        precio = float(row[3].replace('$', '').replace(',', '')) if len(row) > 3 else 0
                        stock = int(row[4]) if len(row) > 4 else 0
                        
                        if nombre and precio > 0:
                            productos_importados.append({
                                'codigo': codigo,
                                'codigo_barras': codigo_barras,
                                'nombre': nombre,
                                'precio': precio,
                                'stock': stock
                            })
            
            if not productos_importados:
                messagebox.showwarning("Sin datos", "No se encontraron productos válidos en el archivo")
                return
            
            # Confirmar importación
            if messagebox.askyesno(
                "Confirmar importación",
                f"Se encontraron {len(productos_importados)} productos.\n\n¿Deseas importarlos?"
            ):
                # Importar productos
                exitosos = 0
                errores = 0
                
                for producto in productos_importados:
                    try:
                        agregar_producto(
                            nombre=producto['nombre'],
                            precio=producto['precio'],
                            stock=producto['stock'],
                            codigo=producto['codigo'] if producto['codigo'] else None,
                            codigo_barras=producto['codigo_barras'] if producto['codigo_barras'] else None
                        )
                        exitosos += 1
                    except Exception as e:
                        print(f"Error al importar {producto['nombre']}: {e}")
                        errores += 1
                
                messagebox.showinfo(
                    "Importación completada",
                    f"Productos importados exitosamente: {exitosos}\n"
                    f"Errores: {errores}\n\n"
                    f"Actualizando lista..."
                )
                
                dialog.destroy()
                self.cargar_tabla_productos()
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar el archivo:\n{str(e)}")

    def abrir_formulario_nuevo(self):
        """Abre la ventana para crear un nuevo producto."""
        def handle_creacion():
            self.cargar_tabla_productos()

        # Usa el formulario mejorado con código de barras
        NuevoProductoFormMejorado(self, on_create=handle_creacion)
    
    def importar_desde_excel(self):
        """Importar productos desde archivo Excel"""
        from tkinter import filedialog
        
        # Abrir diálogo para seleccionar archivo
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[
                ("Archivos Excel", "*.xlsx *.xls"),
                ("Todos los archivos", "*.*")
            ]
        )
        
        if not filename:
            return
        
        try:
            import pandas as pd
            
            # Leer archivo Excel
            df = pd.read_excel(filename)
            
            # Mostrar ventana de vista previa
            self.mostrar_vista_previa_importacion(df, filename)
            
        except ImportError:
            messagebox.showerror(
                "Librería no instalada",
                "Para importar archivos Excel necesitas instalar pandas y openpyxl.\n\n"
                "Ejecuta en la terminal:\n"
                "pip install pandas openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{str(e)}")
    
    def mostrar_vista_previa_importacion(self, df, filename):
        """Mostrar vista previa de los datos a importar"""
        # Crear ventana de vista previa
        preview_window = ctk.CTkToplevel(self)
        preview_window.title("Vista Previa - Importar Productos")
        preview_window.geometry("900x600")
        preview_window.transient(self)
        preview_window.grab_set()
        
        # Centrar ventana
        preview_window.update_idletasks()
        x = (preview_window.winfo_screenwidth() // 2) - (900 // 2)
        y = (preview_window.winfo_screenheight() // 2) - (600 // 2)
        preview_window.geometry(f"900x600+{x}+{y}")
        
        # Contenido
        main_frame = ctk.CTkFrame(preview_window, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        ctk.CTkLabel(
            main_frame,
            text="📥 Vista Previa de Importación",
            font=("Segoe UI", 18, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 10))
        
        ctk.CTkLabel(
            main_frame,
            text=f"Archivo: {filename.split('/')[-1]}",
            font=("Segoe UI", 11),
            text_color="#666666"
        ).pack(pady=(0, 20))
        
        # Información
        info_frame = ctk.CTkFrame(main_frame, fg_color="#E3F2FD", corner_radius=10)
        info_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(
            info_frame,
            text=f"ℹ️ Se encontraron {len(df)} productos en el archivo",
            font=("Segoe UI", 12),
            text_color="#2196F3"
        ).pack(padx=15, pady=12)
        
        # Instrucciones
        inst_frame = ctk.CTkFrame(main_frame, fg_color="#FFF9C4", corner_radius=10)
        inst_frame.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(
            inst_frame,
            text="📋 El archivo debe tener las columnas: nombre, precio, stock, codigo (opcional), codigo_barras (opcional)",
            font=("Segoe UI", 10),
            text_color="#F57C00",
            wraplength=800
        ).pack(padx=15, pady=10)
        
        # Vista previa de datos
        preview_frame = ctk.CTkScrollableFrame(main_frame, height=300)
        preview_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Mostrar primeras 10 filas
        for i, row in df.head(10).iterrows():
            row_frame = ctk.CTkFrame(preview_frame, fg_color="#F5F5F5", corner_radius=8)
            row_frame.pack(fill="x", pady=3)
            
            row_text = f"#{i+1}: "
            if 'nombre' in row:
                row_text += f"{row['nombre']} - "
            if 'precio' in row:
                row_text += f"${row['precio']} - "
            if 'stock' in row:
                row_text += f"Stock: {row['stock']}"
            
            ctk.CTkLabel(
                row_frame,
                text=row_text,
                font=("Segoe UI", 10),
                text_color="#333333",
                anchor="w"
            ).pack(padx=10, pady=8, anchor="w")
        
        if len(df) > 10:
            ctk.CTkLabel(
                preview_frame,
                text=f"... y {len(df) - 10} productos más",
                font=("Segoe UI", 10, "italic"),
                text_color="#999999"
            ).pack(pady=10)
        
        # Botones
        buttons_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        buttons_frame.pack(fill="x")
        
        ctk.CTkButton(
            buttons_frame,
            text="✅ Importar Productos",
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=45,
            font=("Segoe UI", 12, "bold"),
            command=lambda: self.confirmar_importacion(df, preview_window)
        ).pack(side="left", expand=True, fill="x", padx=(0, 5))
        
        ctk.CTkButton(
            buttons_frame,
            text="Cancelar",
            fg_color="#E0E0E0",
            text_color="#666666",
            hover_color="#D0D0D0",
            height=45,
            font=("Segoe UI", 12),
            command=preview_window.destroy
        ).pack(side="left", expand=True, fill="x", padx=(5, 0))
    
    def confirmar_importacion(self, df, window):
        """Confirmar e importar productos a la base de datos"""
        try:
            from database.db import crear_conexion
            
            conn = crear_conexion()
            if not conn:
                messagebox.showerror("Error", "No se pudo conectar a la base de datos")
                return
            
            cursor = conn.cursor()
            productos_importados = 0
            productos_error = 0
            
            for index, row in df.iterrows():
                try:
                    # Validar campos requeridos
                    if 'nombre' not in row or 'precio' not in row or 'stock' not in row:
                        productos_error += 1
                        continue
                    
                    nombre = str(row['nombre'])
                    precio = float(row['precio'])
                    stock = int(row['stock'])
                    codigo = str(row.get('codigo', '')) if 'codigo' in row else None
                    codigo_barras = str(row.get('codigo_barras', '')) if 'codigo_barras' in row else None
                    
                    # Insertar producto
                    if codigo and codigo_barras:
                        cursor.execute(
                            "INSERT INTO productos (nombre, precio, stock, codigo, codigo_barras) VALUES (%s, %s, %s, %s, %s)",
                            (nombre, precio, stock, codigo, codigo_barras)
                        )
                    elif codigo:
                        cursor.execute(
                            "INSERT INTO productos (nombre, precio, stock, codigo) VALUES (%s, %s, %s, %s)",
                            (nombre, precio, stock, codigo)
                        )
                    else:
                        cursor.execute(
                            "INSERT INTO productos (nombre, precio, stock) VALUES (%s, %s, %s)",
                            (nombre, precio, stock)
                        )
                    
                    productos_importados += 1
                    
                except Exception as e:
                    print(f"Error al importar producto {index}: {e}")
                    productos_error += 1
            
            conn.commit()
            conn.close()
            
            # Cerrar ventana
            window.destroy()
            
            # Mostrar resultado
            messagebox.showinfo(
                "Importación completada",
                f"✅ Productos importados: {productos_importados}\n"
                f"❌ Productos con error: {productos_error}\n\n"
                f"Total procesados: {len(df)}"
            )
            
            # Recargar tabla
            self.cargar_tabla_productos()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error durante la importación:\n{str(e)}")

    def editar_producto(self, producto):
        """Editar información de un producto"""
        # Crear ventana de edición
        edit_window = ctk.CTkToplevel(self)
        edit_window.title(f"Editar Producto")
        edit_window.geometry("500x650")
        edit_window.transient(self)
        edit_window.grab_set()
        
        # Centrar ventana
        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (edit_window.winfo_screenheight() // 2) - (650 // 2)
        edit_window.geometry(f"500x650+{x}+{y}")
        
        # Contenido
        main_frame = ctk.CTkFrame(edit_window, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            main_frame,
            text="✏️ Editar Producto",
            font=("Segoe UI", 20, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 20))
        
        # Formulario
        form_frame = ctk.CTkScrollableFrame(main_frame, fg_color="transparent")
        form_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Código
        ctk.CTkLabel(form_frame, text="Código:", font=("Segoe UI", 13, "bold"), text_color="#666666", anchor="w").pack(anchor="w", pady=(0, 5))
        codigo_entry = ctk.CTkEntry(form_frame, placeholder_text="Código del producto", height=40, font=("Segoe UI", 13))
        codigo_entry.pack(fill="x", pady=(0, 15))
        codigo_entry.insert(0, str(producto.get("codigo", "")))
        
        # Nombre
        ctk.CTkLabel(form_frame, text="Nombre del Producto:", font=("Segoe UI", 13, "bold"), text_color="#666666", anchor="w").pack(anchor="w", pady=(0, 5))
        nombre_entry = ctk.CTkEntry(form_frame, placeholder_text="Nombre", height=40, font=("Segoe UI", 13))
        nombre_entry.pack(fill="x", pady=(0, 15))
        nombre_entry.insert(0, str(producto.get("nombre", "")))
        
        # Descripción
        ctk.CTkLabel(form_frame, text="Descripción:", font=("Segoe UI", 13, "bold"), text_color="#666666", anchor="w").pack(anchor="w", pady=(0, 5))
        descripcion_entry = ctk.CTkTextbox(form_frame, height=80, font=("Segoe UI", 13))
        descripcion_entry.pack(fill="x", pady=(0, 15))
        if producto.get("descripcion"):
            descripcion_entry.insert("1.0", producto.get("descripcion"))
        
        # Precio
        ctk.CTkLabel(form_frame, text="Precio:", font=("Segoe UI", 13, "bold"), text_color="#666666", anchor="w").pack(anchor="w", pady=(0, 5))
        precio_entry = ctk.CTkEntry(form_frame, placeholder_text="$0.00", height=40, font=("Segoe UI", 13))
        precio_entry.pack(fill="x", pady=(0, 15))
        precio_entry.insert(0, str(producto.get("precio", 0)))
        
        # Stock actual (solo lectura)
        ctk.CTkLabel(form_frame, text="Stock Actual:", font=("Segoe UI", 13, "bold"), text_color="#666666", anchor="w").pack(anchor="w", pady=(0, 5))
        stock_frame = ctk.CTkFrame(form_frame, fg_color="#E8F5E9", corner_radius=10)
        stock_frame.pack(fill="x", pady=(0, 15))
        
        stock_actual = int(producto.get("stock", 0))
        color_stock = "#E53935" if stock_actual < 10 else "#FF9800" if stock_actual < 20 else "#4CAF50"
        
        ctk.CTkLabel(
            stock_frame,
            text=f"{stock_actual} unidades",
            font=("Segoe UI", 18, "bold"),
            text_color=color_stock
        ).pack(padx=15, pady=12)
        
        ctk.CTkLabel(
            form_frame,
            text="💡 Para ajustar el stock, usa el botón '⚙ Ajustar' en la tabla",
            font=("Segoe UI", 11),
            text_color="#666666"
        ).pack(anchor="w", pady=(0, 15))
        
        # Botones
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x")
        
        def guardar_cambios():
            codigo = codigo_entry.get().strip()
            nombre = nombre_entry.get().strip()
            descripcion = descripcion_entry.get("1.0", "end-1c").strip()
            precio = precio_entry.get().strip()
            
            if not nombre or not precio:
                messagebox.showwarning("Datos incompletos", "Nombre y precio son obligatorios")
                return
            
            try:
                precio_float = float(precio)
                
                if precio_float < 0:
                    messagebox.showwarning("Error", "El precio no puede ser negativo")
                    return
                
                # Usar controlador
                exito = editar_producto(
                    id_producto=producto['id_producto'],
                    nombre=nombre,
                    descripcion=descripcion,
                    precio=precio_float,
                    stock=int(producto.get("stock", 0)),
                    codigo=codigo,
                    codigo_barras=producto.get("codigo_barras")
                )
                
                if exito:
                    messagebox.showinfo("Éxito", f"✅ Producto '{nombre}' actualizado correctamente")
                    edit_window.destroy()
                    self.cargar_tabla_productos()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el producto")
                
            except ValueError:
                messagebox.showerror("Error", "El precio debe ser un número válido")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo actualizar el producto:\n{str(e)}")
        
        ctk.CTkButton(
            btn_frame,
            text="💾 Guardar Cambios",
            fg_color="#4CAF50",
            hover_color="#45a049",
            height=45,
            font=("Segoe UI", 13, "bold"),
            command=guardar_cambios
        ).pack(side="left", expand=True, fill="x", padx=(0, 5))
        
        ctk.CTkButton(
            btn_frame,
            text="Cancelar",
            fg_color="#757575",
            hover_color="#616161",
            height=45,
            font=("Segoe UI", 13, "bold"),
            command=edit_window.destroy
        ).pack(side="left", expand=True, fill="x", padx=(5, 0))

    def eliminar_producto(self, producto):
        """Eliminar un producto con confirmación"""
        from tkinter import messagebox
        
        # Confirmar eliminación
        respuesta = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Estás seguro de que deseas eliminar el producto?\n\n"
            f"Nombre: {producto.get('nombre', 'N/A')}\n"
            f"Código: {producto.get('codigo', 'N/A')}\n\n"
            f"Esta acción no se puede deshacer.",
            icon='warning'
        )
        
        if not respuesta:
            return
        
        try:
            # Usar controlador
            if eliminar_producto(producto.get('id_producto')):
                messagebox.showinfo(
                    "Éxito",
                    f"Producto '{producto.get('nombre', 'N/A')}' eliminado correctamente"
                )
                
                # Recargar lista de productos
                self.cargar_tabla_productos()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el producto")
            
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo eliminar el producto:\n{str(e)}"
            )
    def exportar_excel_pandas(self, dialog):
        """Exportar inventario a Excel usando pandas con formato profesional"""
        try:
            from utils.exportar_pandas import ExportadorPandas
            
            # Usar el exportador de pandas
            exportador = ExportadorPandas()
            filename = exportador.exportar_productos_excel(self.productos)
            
            if filename:
                dialog.destroy()
                messagebox.showinfo(
                    "✅ Exportación Exitosa", 
                    f"Inventario exportado exitosamente a:\n\n📊 {filename}\n\n"
                    f"El archivo incluye:\n"
                    f"• Formato profesional con colores\n"
                    f"• Resumen de totales\n"
                    f"• Columnas ajustadas automáticamente\n"
                    f"• Formato de moneda"
                )
                
                # Preguntar si quiere abrir el archivo
                if messagebox.askyesno("Abrir Archivo", "¿Deseas abrir el archivo Excel ahora?"):
                    import os
                    os.startfile(filename)  # Windows
            else:
                messagebox.showerror("Error", "No se pudo exportar el inventario a Excel")
                
        except ImportError:
            messagebox.showerror(
                "Dependencia Faltante", 
                "Para exportar a Excel necesitas instalar pandas y openpyxl:\n\n"
                "pip install pandas openpyxl\n\n"
                "Mientras tanto, puedes usar la exportación a CSV."
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el inventario: {str(e)}")
    
    def exportar_csv_pandas(self, dialog):
        """Exportar inventario a CSV usando pandas"""
        try:
            from utils.exportar_pandas import ExportadorPandas
            
            exportador = ExportadorPandas()
            filename = exportador.exportar_productos_csv(self.productos)
            
            if filename:
                dialog.destroy()
                messagebox.showinfo(
                    "✅ Exportación Exitosa", 
                    f"Inventario exportado exitosamente a:\n\n📄 {filename}\n\n"
                    f"El archivo CSV es compatible con:\n"
                    f"• Microsoft Excel\n"
                    f"• Google Sheets\n"
                    f"• LibreOffice Calc"
                )
                
                # Preguntar si quiere abrir el archivo
                if messagebox.askyesno("Abrir Archivo", "¿Deseas abrir el archivo CSV ahora?"):
                    import os
                    os.startfile(filename)  # Windows
            else:
                messagebox.showerror("Error", "No se pudo exportar el inventario a CSV")
                
        except ImportError:
            # Fallback a la función CSV básica
            self.exportar_csv(dialog)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el inventario: {str(e)}")
    
    def exportar_html_pandas(self, dialog):
        """Exportar inventario a HTML usando pandas con estilo profesional"""
        try:
            from utils.exportar_pandas import ExportadorPandas
            
            exportador = ExportadorPandas()
            filename = exportador.exportar_productos_html(self.productos)
            
            if filename:
                dialog.destroy()
                messagebox.showinfo(
                    "✅ Exportación Exitosa", 
                    f"Inventario exportado exitosamente a:\n\n📝 {filename}\n\n"
                    f"El archivo HTML incluye:\n"
                    f"• Diseño profesional con colores corporativos\n"
                    f"• Resumen de totales\n"
                    f"• Compatible con navegadores web\n"
                    f"• Se puede abrir en Word"
                )
                
                # Preguntar si quiere abrir el archivo
                if messagebox.askyesno("Abrir Archivo", "¿Deseas abrir el archivo HTML ahora?"):
                    import webbrowser
                    webbrowser.open(filename)
            else:
                messagebox.showerror("Error", "No se pudo exportar el inventario a HTML")
                
        except ImportError:
            # Fallback a la función HTML básica
            self.exportar_html(dialog)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el inventario: {str(e)}")
    
    def exportar_json_pandas(self, dialog):
        """Exportar inventario a JSON usando pandas"""
        try:
            from utils.exportar_pandas import ExportadorPandas
            
            exportador = ExportadorPandas()
            filename = exportador.exportar_productos_json(self.productos)
            
            if filename:
                dialog.destroy()
                messagebox.showinfo(
                    "✅ Exportación Exitosa", 
                    f"Inventario exportado exitosamente a:\n\n📋 {filename}\n\n"
                    f"El archivo JSON es útil para:\n"
                    f"• Integración con otros sistemas\n"
                    f"• Respaldos de datos\n"
                    f"• Desarrollo de aplicaciones\n"
                    f"• APIs y servicios web"
                )
                
                # Preguntar si quiere abrir el archivo
                if messagebox.askyesno("Abrir Archivo", "¿Deseas abrir el archivo JSON ahora?"):
                    import os
                    os.startfile(filename)  # Windows
            else:
                messagebox.showerror("Error", "No se pudo exportar el inventario a JSON")
                
        except ImportError:
            messagebox.showerror(
                "Dependencia Faltante", 
                "Para exportar a JSON necesitas instalar pandas:\n\n"
                "pip install pandas"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el inventario: {str(e)}")